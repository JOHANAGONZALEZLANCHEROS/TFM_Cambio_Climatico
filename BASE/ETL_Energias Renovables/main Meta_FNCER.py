import csv
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def detect_delimiter(file_path, encodings=None):
    """
    Intenta detectar el delimitador leyendo una porción del archivo.
    Retorna el delimitador detectado o None si falla.
    """
    if encodings is None:
        encodings = ['utf-8', 'latin1', 'iso-8859-1']
    
    for enc in encodings:
        try:
            with open(file_path, 'r', encoding=enc, errors='replace') as f:
                sample = f.read(4096)  # lee un fragmento
                f.seek(0)
                # Usa Sniffer para detectar delimitador
                try:
                    dialect = csv.Sniffer().sniff(sample)
                    return dialect.delimiter, enc
                except csv.Error:
                    # No se pudo detectar con este encoding
                    pass
        except Exception:
            pass
    return None, None

def extract(file_path):
    """
    Intenta detectar delimitador y encoding.
    Luego lee el CSV con on_bad_lines='skip', engine='python', etc.
    """
    detected_delim, detected_enc = detect_delimiter(file_path)
    
    if detected_delim is None or detected_enc is None:
        print("[EXTRACT] No se pudo detectar delimitador ni encoding automáticamente.")
        # Opcional: fuerza un delimitador
        detected_delim = ','  # Ajusta a ',' ';' o '\t' según lo que sospeches
        detected_enc = 'utf-8'
    
    print(f"[EXTRACT] Usando delimitador='{detected_delim}' y encoding='{detected_enc}' para leer el CSV.")
    
    try:
        df = pd.read_csv(
            file_path,
            encoding=detected_enc,
            sep=detected_delim,
            quotechar='"',        # Ajusta si tu CSV usa comillas simples
            skipinitialspace=True,
            on_bad_lines='skip',
            engine='python'
        )
        print(f"[EXTRACT] Leído -> {df.shape[0]} filas, {df.shape[1]} columnas.")
        return df
    except Exception as e:
        print(f"[EXTRACT] Error al leer CSV con delimitador '{detected_delim}' y encoding '{detected_enc}': {e}")
        return None

def transform(df):
    """
    - Elimina filas completamente vacías.
    - Convierte 'fecha' a datetime si existe.
    """
    try:
        df = df.dropna(how='all')
        print(f"[TRANSFORM] Filas después de eliminar vacías: {df.shape[0]}")
        
        if 'fecha' in df.columns:
            df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
            print("[TRANSFORM] Columna 'fecha' convertida a datetime.")
        
        return df
    except Exception as e:
        print(f"[TRANSFORM] Error: {e}")
        return df

def load_to_excel(df, output_path):
    """
    Exporta DataFrame a Excel con:
      - Autoajuste de ancho de columnas
      - Estilo de encabezados
    """
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos', index=False)
            ws = writer.sheets['Datos']
            
            # Ajusta ancho de columnas
            for col_idx, col in enumerate(df.columns, 1):
                column_cells = ws[get_column_letter(col_idx)]
                max_length = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                    default=0
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
            
            # Estilo de encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        print(f"[LOAD] Archivo Excel guardado en: {output_path}")
    except Exception as e:
        print(f"[LOAD] Error al guardar Excel: {e}")

if __name__ == "__main__":
    input_file = (
        
        "Meta_FNCER__Incorporar_en_la_matriz_energ_tica_nueva_capacidad_instalada_a_"
        "partir_de_Fuentes_No_Convencionales_de_Energ_a_Renovable_-_FNCER_20250227 (1).csv"
    )
    output_file = "processed_Meta_FNCER_data.xlsx"
    
    df_extracted = extract(input_file)
    if df_extracted is not None and not df_extracted.empty:
        df_transformed = transform(df_extracted)
        if df_transformed is not None and not df_transformed.empty:
            load_to_excel(df_transformed, output_file)
        else:
            print("[INFO] DataFrame transformado está vacío. No se genera archivo.")
    else:
        print("[INFO] No se pudo extraer datos o el DataFrame está vacío.")
